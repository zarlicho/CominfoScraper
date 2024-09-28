from lxml import html
from bs4 import BeautifulSoup
import openpyxl,os,re,requests,threading
from colorama import init,Fore
from InquirerPy import inquirer
from InquirerPy.base.control import Choice

init(autoreset=True)
class Cominfo:
    def __init__(self):
        self.headers = {
            'Accept': '*/*',
            'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Accept-Language': 'en-US,en;q=0.9,id;q=0.8',
            'Origin': 'https://www.kominfo.go.id',
            'Referer': 'https://www.kominfo.go.id/',
            'Sec-CH-UA': '"Google Chrome";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
            'Sec-CH-UA-Mobile': '?0',
            'Sec-CH-UA-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
        }
    
    def storeData(self,filename, data):
        if os.path.exists(filename):
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            start_row = sheet.max_row + 1
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(['Title', 'Body', 'URL', 'URL counter','Image','Date'])
            start_row = 2
        for item in data:
            title = item['title']
            body = item['body']
            url = item['url']
            url_counters = item['url_counters']
            image = item['image']
            date = item['date']
            sheet.append([title, body, url, url_counters[0] if url_counters else '',image,date])
            for counter in url_counters[1:]:
                sheet.append(['', '', '', counter])

            sheet.append(['', '', '', ''])
        workbook.save(filename)
        print(Fore.GREEN + f"Data has been appended to '{filename}'.")

    def extractnews(self,body):
        soup = BeautifulSoup(body, 'html.parser')
        text_content = soup.get_text(separator='\n', strip=True)
        penjelasan_match = re.search(r'Penjelasan:(.*?)(?=Kategori:|$)', text_content, re.DOTALL)
        kategori_match = re.search(r'Kategori:\s*(.*?)(?=Link Counter:|$)', text_content, re.DOTALL)
        link_counter_match = re.search(r'Link Counter:(.*?)$', text_content, re.DOTALL)
        penjelasan = penjelasan_match.group(1).strip() if penjelasan_match else ""
        kategori = kategori_match.group(1).strip() if kategori_match else ""
        link_counter = link_counter_match.group(1).strip() if link_counter_match else ""
        links = [a['href'] for a in soup.find_all('a', href=True)]
        return {
            "penjelasan": penjelasan,
            "kategori": kategori,
            "link_counter": link_counter,
            "links": links
        }

    def getHoax(self,pages):
        response = requests.get(f"https://web.kominfo.go.id/api/v1/contents/category/berita-hoaks?perPage=12&page={pages}&keyword=", headers=self.headers)
        if response.status_code == 200:
            data = response.json()
            for item in data['response']['data']:
                print(item['title'])               
                extraction = self.extractnews(item['body'])
                img = ""
                if "Images" in list(item.keys()):
                    img = item['images'][0]['medium']
        
                new_data = [
                    {
                        'title': item['title'],
                        'body': extraction['penjelasan'],
                        'url': f"https://www.kominfo.go.id/berita/berita-hoaks/detail/{item['slug']}",
                        'url_counters': [line.strip() for line in extraction['link_counter'].strip().split('\n') if line.strip().startswith(('http://', 'https://'))],
                        'date':item['published_at'],
                        'image':img
                    }
                ]
                self.storeData('hoaxData.xlsx', new_data)           
        else:
            return (f"Request failed with status code: {response.status_code}")
        
    def getSatker(self,pages):
        response = requests.get(f"https://web.kominfo.go.id/api/v1/contents/category/berita-kominfo?perPage=12&page={pages}&keyword=", headers=self.headers)
        if response.status_code == 200:
            data = response.json()
            for item in data['response']['data']:
                print(item['title'])     
                penjelasan = ""
                img = ""
                if "Images" in list(item.keys()):
                    img = item['images'][0]['medium']
                for paragraph in [p.get_text(strip=True) for p in BeautifulSoup(item['body'], 'html.parser').find_all('p') if p.get_text(strip=True)]:
                    penjelasan = penjelasan+paragraph + "\n"
                new_data = [
                    {
                        'title': item['title'],
                        'body': penjelasan,
                        'url': f"https://www.kominfo.go.id/berita/berita-kominfo/detail/{item['slug']}",
                        'url_counters':"",
                        'date':item['published_at'],
                        'image':img
                    }
                ]
                self.storeData('satkerData.xlsx', new_data)           
        else:
            return (f"Request failed with status code: {response.status_code}")

class UI:
    def __init__(self):
        self.proceed= False
        self.kominfo = Cominfo()

    def main(self):
        action = inquirer.select(
            message="Select an action:",
            choices=[
                "Isu Hoax",
                "Satker",
                "both",
                Choice(value=None, name="Exit"),
            ],
            default=None,
        ).execute()

        if action == "Isu Hoax":
            page = inquirer.text(
                message="insert jumlah page yang ingin di scrape: ",
                multicolumn_complete=True,
            ).execute()
            for pages in range(int(page.split("-")[0]),int(page.split("-")[1])):
                self.kominfo.getHoax(pages=int(pages))
        elif action == "Satker":
            page = inquirer.text(
                message="insert jumlah page yang ingin di scrape: ",
                multicolumn_complete=True,
            ).execute()
            for pages in range(int(page.split("-")[0]),int(page.split("-")[1])):
                self.kominfo.getSatker(pages=int(pages))
        elif action == "both":
            page = inquirer.text(
                message="insert jumlah page yang ingin di scrape: ",
                multicolumn_complete=True,
            ).execute()
            for pages in range(int(page.split("-")[0]),int(page.split("-")[1])):
                t1 = threading.Thread(target=self.kominfo.getSatker, args=(int(pages),))
                t2 = threading.Thread(target=self.kominfo.getHoax, args=(int(pages),))
                t1.start()
                t2.start()
                t1.join()
                t2.join()

            print(Fore.GREEN + "Done!")
            
if __name__ == "__main__":
    os.system('cls')
    myui = UI()
    myui.main()